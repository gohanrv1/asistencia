import sys
import re
import uuid
import os
from openpyxl import load_workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'backend'))
from main import SessionLocal, Group, Person

def split_name(full):
    full = re.sub(r'\s+', ' ', str(full).strip())
    if not full:
        return '', ''
    parts = full.split()
    if len(parts) == 1:
        return parts[0], ''
    if len(parts) == 2:
        return parts[0], parts[1]
    # Heurística: las dos últimas palabras son apellidos, el resto nombres
    return ' '.join(parts[:-2]), ' '.join(parts[-2:])

def clean_text(value):
    if value is None:
        return ''
    return str(value).strip()

def main():
    db = SessionLocal()
    try:
        wb = load_workbook(r'c:\Users\Gohan\Downloads\LISTAS DOCENTES - META.xlsx', data_only=True)
        created_groups = 0
        created_persons = 0
        assigned = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            group_name = sheet_name.strip().title()

            # Crear grupo
            group = Group(id=str(uuid.uuid4()), name=group_name, description="")
            db.add(group)
            db.flush()
            created_groups += 1
            print(f"[+] Grupo creado: {group_name}")

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                full_name = clean_text(row[0])
                if full_name.lower() == 'total':
                    continue

                cedula = clean_text(row[1])
                if not cedula:
                    continue

                correo = clean_text(row[4])
                telefono = clean_text(row[5])
                institucion = clean_text(row[6])

                nombres, apellidos = split_name(full_name)

                # Buscar persona por cédula para evitar duplicados
                person = db.query(Person).filter(Person.cedula == cedula).first()
                if not person:
                    person = Person(
                        id=str(uuid.uuid4()),
                        nombres=nombres,
                        apellidos=apellidos,
                        cedula=cedula,
                        cargo=institucion or 'Docente Aprendiz',
                        tipo='asistente',
                        correo=correo,
                        celular=telefono
                    )
                    db.add(person)
                    db.flush()
                    created_persons += 1

                # Asignar persona al grupo si no está asignada
                if person not in group.persons:
                    group.persons.append(person)
                    assigned += 1

        db.commit()
        print(f"\nImportacion completa:")
        print(f"   Grupos creados: {created_groups}")
        print(f"   Personas creadas: {created_persons}")
        print(f"   Asignaciones a grupos: {assigned}")
    except Exception as e:
        db.rollback()
        print(f"\nError: {e}")
        raise
    finally:
        db.close()

if __name__ == "__main__":
    main()
