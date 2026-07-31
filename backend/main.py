import sys
if sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

from fastapi import FastAPI, HTTPException, Depends, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy import create_engine, Column, String, Boolean, DateTime, ForeignKey, Text, Table, text
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, Session, relationship
from pydantic import BaseModel, computed_field, Field
from typing import List, Optional
from datetime import datetime
import uuid
import os
import csv
import io
import shutil
import time
import hashlib
import secrets
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Base de datos ─────────────────────────────────────────────────────────────
DATABASE_URL = os.getenv("DATABASE_URL", "sqlite:///./asistencia.db")
engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

# Tabla asociativa muchos a muchos para Grupos y Personas
group_person = Table(
    "group_person",
    Base.metadata,
    Column("group_id", String, ForeignKey("groups.id", ondelete="CASCADE"), primary_key=True),
    Column("person_id", String, ForeignKey("persons.id", ondelete="CASCADE"), primary_key=True)
)

# ── Modelos ORM ───────────────────────────────────────────────────────────────
class Project(Base):
    __tablename__ = "projects"
    id          = Column(String, primary_key=True, default=lambda: str(uuid.uuid4()))
    name        = Column(String, nullable=False)
    description = Column(Text, default="")
    created_at  = Column(DateTime, default=datetime.utcnow)
    events      = relationship("Event",  back_populates="project", cascade="all, delete-orphan")

class Group(Base):
    __tablename__ = "groups"
    id          = Column(String, primary_key=True, default=lambda: str(uuid.uuid4()))
    name        = Column(String, nullable=False)
    description = Column(Text, default="")
    created_at  = Column(DateTime, default=datetime.utcnow)
    persons     = relationship("Person", secondary=group_person, back_populates="groups")

class Person(Base):
    __tablename__ = "persons"
    id         = Column(String, primary_key=True, default=lambda: str(uuid.uuid4()))
    nombres    = Column(String, nullable=False)
    apellidos  = Column(String, nullable=False)
    cedula     = Column(String, nullable=False)
    cargo      = Column(String, nullable=False)
    tipo       = Column(String, nullable=False, default="asistente")
    correo     = Column(String, default="")
    celular    = Column(String, default="")
    groups     = relationship("Group", secondary=group_person, back_populates="persons")

class Event(Base):
    __tablename__ = "events"
    id         = Column(String, primary_key=True, default=lambda: str(uuid.uuid4()))
    name       = Column(String, nullable=False)
    date       = Column(String, nullable=False)
    notes      = Column(Text, default="")
    project_id = Column(String, ForeignKey("projects.id"), nullable=False)
    responsible_id = Column(String, ForeignKey("persons.id"), nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)
    project    = relationship("Project", back_populates="events")
    attendance = relationship("Attendance", back_populates="event", cascade="all, delete-orphan")

class Attendance(Base):
    __tablename__ = "attendance"
    id        = Column(String, primary_key=True, default=lambda: str(uuid.uuid4()))
    event_id  = Column(String, ForeignKey("events.id"), nullable=False)
    person_id = Column(String, nullable=False)
    present   = Column(Boolean, default=False)
    signature = Column(Text, default="")
    updated_at= Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    event     = relationship("Event", back_populates="attendance")

class User(Base):
    __tablename__ = "users"
    id            = Column(String, primary_key=True, default=lambda: str(uuid.uuid4()))
    username      = Column(String, unique=True, nullable=False)
    password_hash = Column(String, nullable=False)
    is_active     = Column(Boolean, default=True)
    role          = Column(String, default="user") # "admin" o "user"
    created_at    = Column(DateTime, default=datetime.utcnow)

class UserSession(Base):
    __tablename__ = "user_sessions"
    token      = Column(String, primary_key=True)
    user_id    = Column(String, ForeignKey("users.id", ondelete="CASCADE"), nullable=False)
    expires_at = Column(DateTime, nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow)

# ── Helpers de Contraseñas ──────────────────────────────────────────────────
def hash_password(password: str) -> str:
    salt = secrets.token_hex(16)
    key = hashlib.pbkdf2_hmac(
        'sha256',
        password.encode('utf-8'),
        salt.encode('utf-8'),
        100000
    )
    return f"pbkdf2_sha256$100000${salt}${key.hex()}"

def verify_password(password: str, hashed: str) -> bool:
    try:
        parts = hashed.split('$')
        if len(parts) != 4 or parts[0] != 'pbkdf2_sha256':
            return False
        iterations = int(parts[1])
        salt = parts[2]
        original_hash = parts[3]
        new_hash = hashlib.pbkdf2_hmac(
            'sha256',
            password.encode('utf-8'),
            salt.encode('utf-8'),
            iterations
        )
        return new_hash.hex() == original_hash
    except Exception:
        return False

# Verificación e inicialización de base de datos
# Usar checkfirst=True para evitar errores con múltiples workers
try:
    print("🔍 Inicializando base de datos...")
    # Crear tablas si no existen (checkfirst=True evita errores si ya existen)
    Base.metadata.create_all(bind=engine, checkfirst=True)

    # Migración: asegurar que la columna 'tipo' exista en persons
    with engine.connect() as conn:
        try:
            conn.execute(text("SELECT tipo FROM persons LIMIT 1"))
        except Exception:
            print("⚠️  Columna 'tipo' no encontrada en persons, agregando...")
            conn.execute(text("ALTER TABLE persons ADD COLUMN tipo TEXT NOT NULL DEFAULT 'asistente'"))
            conn.commit()

    # Verificar que el esquema es correcto (comprobando tablas críticas)
    with engine.connect() as conn:
        conn.execute(text("SELECT cedula FROM persons LIMIT 1"))
        conn.execute(text("SELECT username FROM users LIMIT 1"))
    print("✅ Base de datos OK")
except Exception as e:
    error_msg = str(e)
    if "no such column" in error_msg or "no such table" in error_msg:
        print(f"⚠️ Esquema desactualizado, recreando...")
        try:
            Base.metadata.drop_all(bind=engine)
            Base.metadata.create_all(bind=engine, checkfirst=True)
            print("✅ Base de datos recreada correctamente")
        except Exception as recreate_error:
            print(f"❌ Error al recrear: {str(recreate_error)[:200]}")
    else:
        print(f"✅ Base de datos inicializada")

# Crear el usuario administrador inicial si no existe
try:
    with SessionLocal() as db:
        admin_user = db.query(User).filter(User.username == "admin").first()
        if not admin_user:
            print("👤 Creando usuario administrador inicial (admin)...")
            admin_user = User(
                username="admin",
                password_hash=hash_password("Admin1040*"),
                is_active=True,
                role="admin"
            )
            db.add(admin_user)
            db.commit()
            print("👤 Usuario admin creado correctamente.")
except Exception as admin_err:
    print(f"⚠️ Error al crear usuario admin: {str(admin_err)}")


# ── App FastAPI ───────────────────────────────────────────────────────────────
app = FastAPI(
    title="Asistencia API", 
    version="1.0.0",
    docs_url="/docs",
    redoc_url="/redoc"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Startup Event ─────────────────────────────────────────────────────────────
@app.on_event("startup")
async def startup_event():
    print("🚀 Aplicación FastAPI iniciada correctamente")
    print("📊 Endpoints disponibles en /docs")

@app.on_event("shutdown")
async def shutdown_event():
    print("🛑 Cerrando aplicación FastAPI...")

# ── Health Check ──────────────────────────────────────────────────────────────
@app.get("/health")
@app.get("/api/health")
def health_check():
    """Health check endpoint para Easypanel/Docker"""
    return {"status": "ok", "service": "asistencia", "timestamp": datetime.utcnow().isoformat()}

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi import Security

security_scheme = HTTPBearer(auto_error=False)

def get_current_user(
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security_scheme),
    token: Optional[str] = None,
    db: Session = Depends(get_db)
):
    token_str = None
    if credentials:
        token_str = credentials.credentials
    elif token:
        token_str = token
        
    if not token_str:
        raise HTTPException(status_code=401, detail="No autorizado (Falta Token)")
    
    session = db.query(UserSession).filter(UserSession.token == token_str).first()
    if not session:
        raise HTTPException(status_code=401, detail="Sesión inválida o expirada")
        
    if session.expires_at < datetime.utcnow():
        db.delete(session)
        db.commit()
        raise HTTPException(status_code=401, detail="Sesión expirada")
        
    user = db.query(User).filter(User.id == session.user_id).first()
    if not user:
        raise HTTPException(status_code=401, detail="Usuario no encontrado")
    if not user.is_active:
        raise HTTPException(status_code=403, detail="Usuario inactivo")
    return user

def get_current_admin(current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Requiere permisos de administrador")
    return current_user

# ── Auth Schemas ─────────────────────────────────────────────────────────────
class LoginPayload(BaseModel):
    username: str
    password: str

class UserCreate(BaseModel):
    username: str
    password: str
    role: Optional[str] = "user"

class UserStatusUpdate(BaseModel):
    is_active: bool

class UserPasswordUpdate(BaseModel):
    password: str

class UserOutModel(BaseModel):
    id: str
    username: str
    role: str
    is_active: bool
    created_at: datetime
    class Config: from_attributes = True

# ── Utilidad: ruta real del archivo sqlite ──────────────────────────────────
def resolve_sqlite_path():
    # soporta valores como sqlite:///./asistencia.db o sqlite:////absolute/path/asistencia.db
    if DATABASE_URL.startswith('sqlite:///'):
        path = DATABASE_URL.split('sqlite:///')[-1]
    elif DATABASE_URL.startswith('sqlite:'):
        path = DATABASE_URL.split('sqlite:')[-1]
    else:
        path = DATABASE_URL
    # si es relativo, relativizar respecto a este archivo
    if not os.path.isabs(path):
        path = os.path.join(os.path.dirname(__file__), path)
    return os.path.abspath(path)

# ── Schemas Pydantic ──────────────────────────────────────────────────────────
class ProjectCreate(BaseModel):
    id: Optional[str] = None
    name: str
    description: Optional[str] = ""

class ProjectOut(BaseModel):
    id: str
    name: str
    description: str
    created_at: datetime
    class Config: from_attributes = True

class PersonCreate(BaseModel):
    model_config = {"populate_by_name": True}
    id: Optional[str] = None
    nombres: Optional[str] = ""
    apellidos: Optional[str] = ""
    cedula: Optional[str] = ""
    cargo: Optional[str] = ""
    tipo: Optional[str] = "asistente"  # "asistente" o "responsable"
    correo: Optional[str] = ""
    celular: Optional[str] = ""
    project_id: Optional[str] = None
    delete: Optional[bool] = Field(default=False, alias="_delete")

class PersonOut(BaseModel):
    id: str
    nombres: str
    apellidos: str
    cedula: str
    cargo: str
    tipo: str
    correo: str
    celular: str

    @computed_field
    @property
    def name(self) -> str:
        return f"{self.nombres} {self.apellidos}"

    @computed_field
    @property
    def role(self) -> str:
        return self.cargo

    class Config: from_attributes = True

class EventCreate(BaseModel):
    id: Optional[str] = None
    name: str
    date: str
    notes: Optional[str] = ""
    project_id: str
    responsible_id: Optional[str] = None
    group_ids: Optional[List[str]] = []
    person_ids: Optional[List[str]] = []

class EventOut(BaseModel):
    id: str
    name: str
    date: str
    notes: str
    project_id: str
    responsible_id: Optional[str]
    created_at: datetime
    class Config: from_attributes = True

class AttendanceRecord(BaseModel):
    person_id: str
    present: bool
    signature: Optional[str] = None

class AttendanceBulk(BaseModel):
    records: List[AttendanceRecord]

class GroupPersonSync(BaseModel):
    group_id: str
    person_id: str

class GroupCreate(BaseModel):
    id: Optional[str] = None
    name: str
    description: Optional[str] = ""
    _delete: Optional[bool] = False

class GroupOut(BaseModel):
    id: str
    name: str
    description: str
    created_at: datetime
    class Config: from_attributes = True

class GroupWithPersonsOut(BaseModel):
    id: str
    name: str
    description: str
    created_at: datetime
    person_ids: List[str]
    class Config: from_attributes = True

class SyncPayload(BaseModel):
    """Payload de sincronización offline → servidor"""
    projects:      List[ProjectCreate]      = []
    persons:       List[PersonCreate]       = []
    events:        List[EventCreate]        = []
    attendance:    dict                     = {}   # { event_id: { person_id: bool } }
    groups:        List[GroupCreate]        = []
    group_assignments:   List[GroupPersonSync] = []
    group_unassignments: List[GroupPersonSync] = []
    deleted_groups: List[str]                = []

from datetime import timedelta

# ── Endpoints: Autenticación ──────────────────────────────────────────────────
@app.post("/api/auth/login")
def login(data: LoginPayload, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == data.username).first()
    if not user or not verify_password(data.password, user.password_hash):
        raise HTTPException(status_code=400, detail="Usuario o contraseña incorrectos")
    if not user.is_active:
        raise HTTPException(status_code=403, detail="Usuario inactivo")
    
    token = secrets.token_hex(32)
    expires_at = datetime.utcnow() + timedelta(days=7)
    
    session = UserSession(token=token, user_id=user.id, expires_at=expires_at)
    db.add(session)
    db.commit()
    
    return {
        "token": token,
        "username": user.username,
        "role": user.role
    }

@app.post("/api/auth/logout")
def logout(current_user: User = Depends(get_current_user), credentials: Optional[HTTPAuthorizationCredentials] = Depends(security_scheme), db: Session = Depends(get_db)):
    if credentials:
        token = credentials.credentials
        session = db.query(UserSession).filter(UserSession.token == token).first()
        if session:
            db.delete(session)
            db.commit()
    return {"ok": True}

@app.get("/api/auth/me", response_model=UserOutModel)
def get_me(current_user: User = Depends(get_current_user)):
    return current_user

# ── Endpoints: Gestión de Usuarios (Sólo Administradores) ──────────────────────
@app.get("/api/users", response_model=List[UserOutModel])
def list_users(current_admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    return db.query(User).order_by(User.created_at.desc()).all()

@app.post("/api/users", response_model=UserOutModel, status_code=201)
def create_user(data: UserCreate, current_admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    existing = db.query(User).filter(User.username == data.username).first()
    if existing:
        raise HTTPException(status_code=400, detail="El nombre de usuario ya está registrado")
    
    new_user = User(
        username=data.username,
        password_hash=hash_password(data.password),
        role=data.role or "user",
        is_active=True
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    return new_user

@app.put("/api/users/{user_id}/status", response_model=UserOutModel)
def update_user_status(user_id: str, data: UserStatusUpdate, current_admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    if user.id == current_admin.id:
        raise HTTPException(status_code=400, detail="No puedes desactivar tu propio usuario")
        
    user.is_active = data.is_active
    db.commit()
    db.refresh(user)
    return user

@app.put("/api/users/{user_id}/password", response_model=UserOutModel)
def update_user_password(user_id: str, data: UserPasswordUpdate, current_admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
        
    user.password_hash = hash_password(data.password)
    db.commit()
    db.refresh(user)
    return user

# ── Endpoints: Proyectos ──────────────────────────────────────────────────────
@app.get("/api/projects", response_model=List[ProjectOut])
def list_projects(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    return db.query(Project).order_by(Project.created_at.desc()).all()

@app.post("/api/projects", response_model=ProjectOut, status_code=201)
def create_project(data: ProjectCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    proj = Project(id=data.id or str(uuid.uuid4()), name=data.name, description=data.description or "")
    db.add(proj); db.commit(); db.refresh(proj)
    return proj

@app.get("/api/projects/{project_id}", response_model=ProjectOut)
def get_project(project_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    proj = db.query(Project).filter(Project.id == project_id).first()
    if not proj: raise HTTPException(404, "Proyecto no encontrado")
    return proj

@app.put("/api/projects/{project_id}", response_model=ProjectOut)
def update_project(project_id: str, data: ProjectCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    proj = db.query(Project).filter(Project.id == project_id).first()
    if not proj: raise HTTPException(404, "Proyecto no encontrado")
    proj.name = data.name; proj.description = data.description or ""
    db.commit(); db.refresh(proj)
    return proj

@app.delete("/api/projects/{project_id}", status_code=204)
def delete_project(project_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    proj = db.query(Project).filter(Project.id == project_id).first()
    if not proj: raise HTTPException(404, "Proyecto no encontrado")
    db.delete(proj); db.commit()

# ── Endpoints: Personas ───────────────────────────────────────────────────────
@app.get("/api/persons", response_model=List[PersonOut])
def list_all_persons(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    return db.query(Person).all()

@app.post("/api/persons", response_model=PersonOut, status_code=201)
def create_person(data: PersonCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    tipo = data.tipo if data.tipo in ("asistente", "responsable") else "asistente"
    person = db.query(Person).filter(Person.cedula == data.cedula).first()
    if not person:
        person = Person(
            id=data.id or str(uuid.uuid4()),
            nombres=data.nombres,
            apellidos=data.apellidos,
            cedula=data.cedula,
            cargo=data.cargo,
            tipo=tipo,
            correo=data.correo or "",
            celular=data.celular or ""
        )
        db.add(person)
    else:
        person.tipo = tipo
    db.commit(); db.refresh(person)
    return person

@app.delete("/api/persons/{person_id}", status_code=204)
def delete_person(person_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    person = db.query(Person).filter(Person.id == person_id).first()
    if not person: raise HTTPException(404, "Persona no encontrada")
    db.delete(person); db.commit()

# ── Endpoints: Grupos ─────────────────────────────────────────────────────────
@app.get("/api/groups", response_model=List[GroupWithPersonsOut])
def list_groups(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    groups = db.query(Group).order_by(Group.created_at.desc()).all()
    return [{"id": g.id, "name": g.name, "description": g.description or "", "created_at": g.created_at, "person_ids": [p.id for p in g.persons]} for g in groups]

@app.post("/api/groups", response_model=GroupOut, status_code=201)
def create_group(data: GroupCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    group = Group(id=data.id or str(uuid.uuid4()), name=data.name, description=data.description or "")
    db.add(group); db.commit(); db.refresh(group)
    return group

@app.get("/api/groups/{group_id}", response_model=GroupWithPersonsOut)
def get_group(group_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    group = db.query(Group).filter(Group.id == group_id).first()
    if not group: raise HTTPException(404, "Grupo no encontrado")
    return {"id": group.id, "name": group.name, "description": group.description or "", "created_at": group.created_at, "person_ids": [p.id for p in group.persons]}

@app.put("/api/groups/{group_id}", response_model=GroupOut)
def update_group(group_id: str, data: GroupCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    group = db.query(Group).filter(Group.id == group_id).first()
    if not group: raise HTTPException(404, "Grupo no encontrado")
    group.name = data.name; group.description = data.description or ""
    db.commit(); db.refresh(group)
    return group

@app.delete("/api/groups/{group_id}", status_code=204)
def delete_group(group_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    group = db.query(Group).filter(Group.id == group_id).first()
    if not group: raise HTTPException(404, "Grupo no encontrado")
    db.delete(group); db.commit()

@app.post("/api/groups/{group_id}/persons/{person_id}", status_code=200)
def assign_person_to_group(group_id: str, person_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    group = db.query(Group).filter(Group.id == group_id).first()
    person = db.query(Person).filter(Person.id == person_id).first()
    if not group or not person:
        raise HTTPException(404, "Grupo o persona no encontrado")
    if person not in group.persons:
        group.persons.append(person)
        db.commit()
    return {"ok": True}

@app.delete("/api/groups/{group_id}/persons/{person_id}", status_code=204)
def unassign_person_from_group(group_id: str, person_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    group = db.query(Group).filter(Group.id == group_id).first()
    person = db.query(Person).filter(Person.id == person_id).first()
    if not group or not person:
        raise HTTPException(404, "Grupo o persona no encontrado")
    if person in group.persons:
        group.persons.remove(person)
        db.commit()

@app.get("/api/group-persons")
def list_group_persons(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    associations = db.query(group_person).all()
    return [{"group_id": a.group_id, "person_id": a.person_id} for a in associations]

@app.get("/api/persons/template")
def download_csv_template(current_user: User = Depends(get_current_user)):
    csv_content = "nombres,apellidos,cedula,cargo,tipo,correo,celular\nJuan,Perez,12345678,Gerente,asistente,juan@example.com,555-1234\nMaria,Gomez,87654321,Analista,responsable,,555-5678\n"
    output = io.StringIO()
    output.write(csv_content)
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.read().encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=plantilla_personas.csv"}
    )

@app.post("/api/persons/upload-csv")
def upload_csv(file: UploadFile = File(...), current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    try:
        content = file.file.read().decode("utf-8")
        reader = csv.DictReader(io.StringIO(content))
    except Exception as e:
        raise HTTPException(400, f"Error al procesar el archivo CSV: {str(e)}")

    created_count = 0

    for row in reader:
        nombres = row.get("nombres", "").strip()
        apellidos = row.get("apellidos", "").strip()
        cedula = row.get("cedula", "").strip()
        cargo = row.get("cargo", "").strip()
        tipo_raw = row.get("tipo", "").strip().lower()
        tipo = tipo_raw if tipo_raw in ("asistente", "responsable") else "asistente"

        if not nombres or not apellidos or not cedula or not cargo:
            continue

        correo = row.get("correo", "").strip()
        celular = row.get("celular", "").strip()

        person = db.query(Person).filter(Person.cedula == cedula).first()
        if not person:
            person = Person(
                id=str(uuid.uuid4()),
                nombres=nombres,
                apellidos=apellidos,
                cedula=cedula,
                cargo=cargo,
                tipo=tipo,
                correo=correo,
                celular=celular
            )
            db.add(person)
            created_count += 1
        else:
            person.tipo = tipo

    db.commit()
    return {"ok": True, "created": created_count, "assigned": 0}

# ── Endpoints: Eventos ────────────────────────────────────────────────────────
@app.get("/api/events", response_model=List[EventOut])
def list_events(project_id: Optional[str] = None, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    q = db.query(Event)
    if project_id: q = q.filter(Event.project_id == project_id)
    return q.order_by(Event.date.desc()).all()

@app.post("/api/events", response_model=EventOut, status_code=201)
def create_event(data: EventCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not db.query(Project).filter(Project.id == data.project_id).first():
        raise HTTPException(404, "Proyecto no encontrado")
    ev = Event(id=data.id or str(uuid.uuid4()), name=data.name, date=data.date, notes=data.notes or "", project_id=data.project_id, responsible_id=data.responsible_id)
    db.add(ev); db.commit(); db.refresh(ev)

    # Precargar asistencia con personas de los grupos seleccionados
    seen = set()
    group_ids = data.group_ids or []
    for gid in group_ids:
        group = db.query(Group).filter(Group.id == gid).first()
        if group:
            for person in group.persons:
                if person.id not in seen and (person.tipo or "asistente") == "asistente":
                    seen.add(person.id)
                    existing = db.query(Attendance).filter(
                        Attendance.event_id == ev.id,
                        Attendance.person_id == person.id
                    ).first()
                    if not existing:
                        db.add(Attendance(event_id=ev.id, person_id=person.id, present=False, signature=""))
    # Precargar personas individuales seleccionadas
    person_ids = data.person_ids or []
    for pid in person_ids:
        if pid not in seen:
            seen.add(pid)
            existing = db.query(Attendance).filter(
                Attendance.event_id == ev.id,
                Attendance.person_id == pid
            ).first()
            if not existing:
                db.add(Attendance(event_id=ev.id, person_id=pid, present=False, signature=""))
    if seen:
        db.commit(); db.refresh(ev)
    return ev

@app.get("/api/events/{event_id}", response_model=EventOut)
def get_event(event_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ev = db.query(Event).filter(Event.id == event_id).first()
    if not ev: raise HTTPException(404, "Evento no encontrado")
    return ev

@app.put("/api/events/{event_id}", response_model=EventOut)
def update_event(event_id: str, data: EventCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ev = db.query(Event).filter(Event.id == event_id).first()
    if not ev: raise HTTPException(404, "Evento no encontrado")
    if not db.query(Project).filter(Project.id == data.project_id).first():
        raise HTTPException(404, "Proyecto no encontrado")
    ev.name = data.name
    ev.date = data.date
    ev.notes = data.notes or ""
    ev.project_id = data.project_id
    ev.responsible_id = data.responsible_id
    db.commit(); db.refresh(ev)
    return ev

@app.delete("/api/events/{event_id}", status_code=204)
def delete_event(event_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ev = db.query(Event).filter(Event.id == event_id).first()
    if not ev: raise HTTPException(404, "Evento no encontrado")
    db.delete(ev); db.commit()

# ── Endpoints: Asistencia ─────────────────────────────────────────────────────
@app.get("/api/events/{event_id}/attendance")
def get_attendance(event_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    records = db.query(Attendance).filter(Attendance.event_id == event_id).all()
    # return present status and signature if exists
    out = {}
    for r in records:
        out[r.person_id] = {"present": r.present, "signature": r.signature if getattr(r, 'signature', None) else None}
    return out

@app.post("/api/events/{event_id}/attendance")
def save_attendance(event_id: str, data: AttendanceBulk, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not db.query(Event).filter(Event.id == event_id).first():
        raise HTTPException(404, "Evento no encontrado")
    for rec in data.records:
        existing = db.query(Attendance).filter(
            Attendance.event_id == event_id,
            Attendance.person_id == rec.person_id
        ).first()
        if existing:
            existing.present = rec.present
            existing.signature = rec.signature or existing.signature
            existing.updated_at = datetime.utcnow()
        else:
            db.add(Attendance(event_id=event_id, person_id=rec.person_id, present=rec.present, signature=rec.signature or ""))
    db.commit()
    return {"ok": True}

@app.delete("/api/events/{event_id}/attendance/{person_id}", status_code=204)
def delete_attendance(event_id: str, person_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not db.query(Event).filter(Event.id == event_id).first():
        raise HTTPException(404, "Evento no encontrado")
    record = db.query(Attendance).filter(
        Attendance.event_id == event_id,
        Attendance.person_id == person_id
    ).first()
    if record:
        db.delete(record); db.commit()

# ── Endpoint: Sincronización offline ─────────────────────────────────────────
@app.post("/api/sync")
def sync_offline(payload: SyncPayload, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """
    Recibe todos los cambios pendientes del cliente offline
    y los aplica en orden: proyectos → personas → eventos → asistencia.
    Usa upsert (insertar si no existe, ignorar si ya existe).
    """
    synced = {"projects": 0, "persons": 0, "events": 0, "attendance": 0, "groups": 0, "group_assignments": 0, "group_unassignments": 0, "deleted_groups": 0}

    for p in payload.projects:
        if not db.query(Project).filter(Project.id == p.id).first():
            db.add(Project(id=p.id, name=p.name, description=p.description or ""))
            synced["projects"] += 1

    db.flush()

    for p in payload.persons:
        person = db.query(Person).filter(Person.id == p.id).first()
        if p.delete:
            if person:
                db.delete(person)
                synced["persons"] = synced.get("persons", 0) + 1
            continue
        tipo = p.tipo if p.tipo in ("asistente", "responsable") else "asistente"
        if not person:
            person = Person(
                id=p.id,
                nombres=p.nombres or "",
                apellidos=p.apellidos or "",
                cedula=p.cedula or "",
                cargo=p.cargo or "",
                tipo=tipo,
                correo=p.correo or "",
                celular=p.celular or ""
            )
            db.add(person)
            synced["persons"] += 1
        else:
            person.tipo = tipo

    db.flush()

    for e in payload.events:
        if not db.query(Event).filter(Event.id == e.id).first():
            if db.query(Project).filter(Project.id == e.project_id).first():
                db.add(Event(id=e.id, name=e.name, date=e.date, notes=e.notes or "", project_id=e.project_id, responsible_id=e.responsible_id))
                synced["events"] += 1
                # Precargar asistencia desde grupos/personas si vienen
                seen = set()
                if e.group_ids:
                    for gid in e.group_ids:
                        group = db.query(Group).filter(Group.id == gid).first()
                        if group:
                            for person in group.persons:
                                if person.id not in seen and (person.tipo or "asistente") == "asistente":
                                    seen.add(person.id)
                                    existing = db.query(Attendance).filter(
                                        Attendance.event_id == e.id,
                                        Attendance.person_id == person.id
                                    ).first()
                                    if not existing:
                                        db.add(Attendance(event_id=e.id, person_id=person.id, present=False, signature=""))
                if e.person_ids:
                    for pid in e.person_ids:
                        if pid not in seen:
                            seen.add(pid)
                            existing = db.query(Attendance).filter(
                                Attendance.event_id == e.id,
                                Attendance.person_id == pid
                            ).first()
                            if not existing:
                                db.add(Attendance(event_id=e.id, person_id=pid, present=False, signature=""))

    db.flush()

    # Eliminar grupos marcados para borrado
    for gid in payload.deleted_groups:
        existing_group = db.query(Group).filter(Group.id == gid).first()
        if existing_group:
            db.delete(existing_group)
            synced["deleted_groups"] = synced.get("deleted_groups", 0) + 1

    for g in payload.groups:
        existing_group = db.query(Group).filter(Group.id == g.id).first()
        if not existing_group:
            db.add(Group(id=g.id, name=g.name, description=g.description or ""))
            synced["groups"] = synced.get("groups", 0) + 1
        else:
            existing_group.name = g.name
            existing_group.description = g.description or ""

    db.flush()

    for assoc in payload.group_assignments:
        group = db.query(Group).filter(Group.id == assoc.group_id).first()
        person = db.query(Person).filter(Person.id == assoc.person_id).first()
        if group and person and person not in group.persons:
            group.persons.append(person)
            synced["group_assignments"] = synced.get("group_assignments", 0) + 1

    for assoc in payload.group_unassignments:
        group = db.query(Group).filter(Group.id == assoc.group_id).first()
        person = db.query(Person).filter(Person.id == assoc.person_id).first()
        if group and person and person in group.persons:
            group.persons.remove(person)
            synced["group_unassignments"] = synced.get("group_unassignments", 0) + 1

    db.flush()

    db.flush()

    for event_id, records in payload.attendance.items():
        for person_id, present_val in records.items():
            # support either boolean or dict { present: bool, signature: str, _delete: bool }
            if isinstance(present_val, dict):
                present = present_val.get('present', False)
                signature = present_val.get('signature')
                delete = present_val.get('_delete', False)
            else:
                present = bool(present_val)
                signature = None
                delete = False

            existing = db.query(Attendance).filter(
                Attendance.event_id == event_id,
                Attendance.person_id == person_id
            ).first()
            if delete:
                if existing:
                    db.delete(existing)
            elif existing:
                existing.present = present
                if signature:
                    existing.signature = signature
            else:
                db.add(Attendance(event_id=event_id, person_id=person_id, present=present, signature=signature or ""))
            synced["attendance"] += 1

    db.commit()
    return {"ok": True, "synced": synced}


# ── Helpers para reportes Excel ───────────────────────────────────────────────
def _style_header(ws, row_idx):
    header_fill = PatternFill(start_color="0C447C", end_color="0C447C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[row_idx]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border


def _auto_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val_len = len(str(cell.value)) if cell.value is not None else 0
                if val_len > max_length:
                    max_length = val_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 50)


def _build_excel_report(db: Session) -> io.BytesIO:
    wb = Workbook()
    wb.remove(wb.active)

    projects = db.query(Project).all()
    persons = db.query(Person).all()
    groups = db.query(Group).all()
    events = db.query(Event).all()
    total_att = db.query(Attendance).count()
    total_present = db.query(Attendance).filter(Attendance.present == True).count()
    attendance_pct = round((total_present / total_att * 100), 1) if total_att else 0

    # ── Hoja 1: Resumen Ejecutivo ─────────────────────────────────────────────
    ws_summary = wb.create_sheet("📊 Resumen Ejecutivo")
    ws_summary.append(["INFORME CONSORCIO META TIC - ASISTENCIA"])
    ws_summary.append(["Generado el", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")])
    ws_summary.append([])
    ws_summary.append(["INDICADORES CLAVE (KPIs)"])
    ws_summary.append(["Métrica", "Valor"])
    ws_summary.append(["Total de proyectos", len(projects)])
    ws_summary.append(["Total de grupos", len(groups)])
    ws_summary.append(["Total de personas registradas", len(persons)])
    ws_summary.append(["Total de eventos realizados", len(events)])
    ws_summary.append(["Total de registros de asistencia", total_att])
    ws_summary.append(["Total de presentes", total_present])
    ws_summary.append(["Total de ausentes", total_att - total_present])
    ws_summary.append(["Porcentaje de asistencia global", f"{attendance_pct}%"])
    _style_header(ws_summary, 5)
    ws_summary.merge_cells("A1:B1")
    ws_summary["A1"].font = Font(size=16, bold=True, color="0C447C")
    ws_summary["A1"].alignment = Alignment(horizontal="center")
    ws_summary["A2"].font = Font(italic=True, color="666666")
    ws_summary.merge_cells("A4:B4")
    ws_summary["A4"].font = Font(size=12, bold=True, color="0C447C")
    ws_summary["A4"].alignment = Alignment(horizontal="center")
    _auto_width(ws_summary)

    # ── Hoja 2: ⚠️ Eventos Críticos (Requieren Atención) ─────────────────────
    ws_critical = wb.create_sheet("⚠️ Eventos Críticos")
    ws_critical.append(["EVENTOS CON BAJA ASISTENCIA (< 70%)"])
    ws_critical.append([])
    ws_critical.append([
        "Evento", "Fecha", "Proyecto", "Responsable", "Esperados",
        "Presentes", "Ausentes", "% Asistencia", "Estado"
    ])
    _style_header(ws_critical, 3)
    ws_critical.merge_cells("A1:I1")
    ws_critical["A1"].font = Font(size=14, bold=True, color="C00000")
    ws_critical["A1"].alignment = Alignment(horizontal="center")

    critical_list = []
    for ev in events:
        expected = db.query(Attendance).filter(Attendance.event_id == ev.id).count()
        if expected == 0:
            continue
        present = db.query(Attendance).filter(Attendance.event_id == ev.id, Attendance.present == True).count()
        pct = round((present / expected * 100), 1)
        if pct < 70:
            resp = ""
            if ev.responsible_id:
                r = db.query(Person).filter(Person.id == ev.responsible_id).first()
                if r: resp = f"{r.nombres} {r.apellidos}".strip()
            critical_list.append({
                "name": ev.name,
                "date": ev.date,
                "project": ev.project.name if ev.project else "",
                "responsible": resp,
                "expected": expected,
                "present": present,
                "absent": expected - present,
                "pct": pct,
                "status": "🔴 CRÍTICO" if pct < 50 else "🟡 BAJO"
            })
    
    critical_list = sorted(critical_list, key=lambda x: x['pct'])
    for item in critical_list:
        ws_critical.append([
            item["name"], item["date"], item["project"], item["responsible"],
            item["expected"], item["present"], item["absent"], f"{item['pct']}%", item["status"]
        ])
    
    if not critical_list:
        ws_critical.append(["No hay eventos con asistencia baja. ¡Excelente trabajo!"])
    _auto_width(ws_critical)

    # ── Hoja 3: 🏆 Ranking de Proyectos ──────────────────────────────────────
    ws_ranking = wb.create_sheet("🏆 Ranking de Proyectos")
    ws_ranking.append(["RANKING DE PROYECTOS POR % DE ASISTENCIA"])
    ws_ranking.append([])
    ws_ranking.append([
        "Posición", "Proyecto", "Descripción", "# Eventos", "Participantes Únicos",
        "Esperados Total", "Presentes Total", "% Asistencia", "Clasificación"
    ])
    _style_header(ws_ranking, 3)
    ws_ranking.merge_cells("A1:I1")
    ws_ranking["A1"].font = Font(size=14, bold=True, color="0C7C0C")
    ws_ranking["A1"].alignment = Alignment(horizontal="center")

    project_ranking = []
    for proj in projects:
        proj_events = proj.events
        if not proj_events:
            continue
        
        # Participantes únicos (personas que han asistido a al menos un evento del proyecto)
        unique_persons = set()
        for ev in proj_events:
            for att in db.query(Attendance).filter(Attendance.event_id == ev.id).all():
                unique_persons.add(att.person_id)
        
        total_present_proj = 0
        total_expected_proj = 0
        for ev in proj_events:
            expected_ev = db.query(Attendance).filter(Attendance.event_id == ev.id).count()
            present_ev = db.query(Attendance).filter(Attendance.event_id == ev.id, Attendance.present == True).count()
            total_present_proj += present_ev
            total_expected_proj += expected_ev
        
        if total_expected_proj > 0:
            pct_proj = round((total_present_proj / total_expected_proj * 100), 1)
            classification = "🥇 Excelente" if pct_proj >= 90 else "🥈 Bueno" if pct_proj >= 75 else "🥉 Regular" if pct_proj >= 60 else "❌ Bajo"
            project_ranking.append({
                "name": proj.name,
                "desc": proj.description or "-",
                "events": len(proj_events),
                "unique_persons": len(unique_persons),
                "expected": total_expected_proj,
                "present": total_present_proj,
                "pct": pct_proj,
                "classification": classification
            })
    
    project_ranking = sorted(project_ranking, key=lambda x: x['pct'], reverse=True)
    for i, item in enumerate(project_ranking, 1):
        ws_ranking.append([
            i, item["name"], item["desc"], item["events"], item["unique_persons"],
            item["expected"], item["present"], f"{item['pct']}%", item["classification"]
        ])
    _auto_width(ws_ranking)

    # ── Hoja 4: 👥 Análisis por Grupos ───────────────────────────────────────
    ws_groups_analysis = wb.create_sheet("👥 Análisis por Grupos")
    ws_groups_analysis.append(["DESEMPEÑO DE ASISTENCIA POR GRUPOS"])
    ws_groups_analysis.append([])
    ws_groups_analysis.append([
        "Grupo", "Descripción", "# Miembros", "Registros Totales",
        "Presentes", "Ausentes", "% Asistencia", "Evaluación"
    ])
    _style_header(ws_groups_analysis, 3)
    ws_groups_analysis.merge_cells("A1:H1")
    ws_groups_analysis["A1"].font = Font(size=14, bold=True, color="0C447C")
    ws_groups_analysis["A1"].alignment = Alignment(horizontal="center")

    group_stats = []
    all_groups = db.query(Group).all()
    for g in all_groups:
        if not g.persons:
            continue
        total_expected_group = 0
        total_present_group = 0
        for p in g.persons:
            att_person = db.query(Attendance).filter(Attendance.person_id == p.id).count()
            pres_person = db.query(Attendance).filter(Attendance.person_id == p.id, Attendance.present == True).count()
            total_expected_group += att_person
            total_present_group += pres_person
        
        if total_expected_group > 0:
            pct_group = round((total_present_group / total_expected_group * 100), 1)
            evaluation = "⭐⭐⭐ Excelente" if pct_group >= 90 else "⭐⭐ Bueno" if pct_group >= 75 else "⭐ Regular" if pct_group >= 60 else "⚠️ Requiere atención"
            group_stats.append({
                "name": g.name,
                "desc": g.description or "-",
                "members": len(g.persons),
                "expected": total_expected_group,
                "present": total_present_group,
                "absent": total_expected_group - total_present_group,
                "pct": pct_group,
                "evaluation": evaluation
            })
    
    group_stats = sorted(group_stats, key=lambda x: x['pct'], reverse=True)
    for item in group_stats:
        ws_groups_analysis.append([
            item["name"], item["desc"], item["members"], item["expected"],
            item["present"], item["absent"], f"{item['pct']}%", item["evaluation"]
        ])
    
    if not group_stats:
        ws_groups_analysis.append(["No hay datos de grupos con asistencia registrada."])
    _auto_width(ws_groups_analysis)

    # ── Hoja 5: 📋 Listado de Todos los Eventos ──────────────────────────────
    ws_events = wb.create_sheet("📋 Todos los Eventos")
    ws_events.append([
        "Evento", "Fecha", "Proyecto", "Responsable", "Asistentes Esperados",
        "Presentes", "Ausentes", "% Asistencia", "% con Firma"
    ])
    _style_header(ws_events, 1)

    events_sorted = sorted(events, key=lambda e: e.date, reverse=True)
    for ev in events_sorted:
        proj = ev.project
        responsible = ""
        if ev.responsible_id:
            resp = db.query(Person).filter(Person.id == ev.responsible_id).first()
            if resp:
                responsible = f"{resp.nombres} {resp.apellidos}".strip()

        expected = db.query(Attendance).filter(Attendance.event_id == ev.id).count()
        present = db.query(Attendance).filter(Attendance.event_id == ev.id, Attendance.present == True).count()
        absent = expected - present
        pct = round((present / expected * 100), 1) if expected else 0
        signed = db.query(Attendance).filter(
            Attendance.event_id == ev.id,
            Attendance.present == True,
            Attendance.signature != ""
        ).count()
        sign_pct = f"{round((signed / present * 100), 1)}%" if present else "N/A"

        ws_events.append([
            ev.name,
            ev.date,
            proj.name if proj else "",
            responsible,
            expected,
            present,
            absent if absent >= 0 else 0,
            f"{pct}%",
            sign_pct
        ])
    _auto_width(ws_events)

    # ── Hoja 6: 🎯 Ranking de Personas ───────────────────────────────────────
    ws_persons = wb.create_sheet("🎯 Ranking de Personas")
    ws_persons.append(["RANKING DE PERSONAS POR ASISTENCIA"])
    ws_persons.append([])
    ws_persons.append([
        "Posición", "Nombres", "Apellidos", "Cédula", "Cargo",
        "Eventos Asignados", "Presentes", "Ausentes", "% Asistencia", "Evaluación"
    ])
    _style_header(ws_persons, 3)
    ws_persons.merge_cells("A1:J1")
    ws_persons["A1"].font = Font(size=14, bold=True, color="0C447C")
    ws_persons["A1"].alignment = Alignment(horizontal="center")

    person_stats = []
    for p in persons:
        att_count = db.query(Attendance).filter(Attendance.person_id == p.id).count()
        if att_count == 0:
            continue
        pres_count = db.query(Attendance).filter(Attendance.person_id == p.id, Attendance.present == True).count()
        pct_person = round((pres_count / att_count * 100), 1)
        evaluation = "🌟 Excelente" if pct_person >= 90 else "👍 Bueno" if pct_person >= 75 else "🔔 Regular" if pct_person >= 60 else "⚠️ Bajo"
        person_stats.append({
            "nombres": p.nombres,
            "apellidos": p.apellidos,
            "cedula": p.cedula,
            "cargo": p.cargo,
            "expected": att_count,
            "present": pres_count,
            "absent": att_count - pres_count,
            "pct": pct_person,
            "evaluation": evaluation
        })
    
    person_stats = sorted(person_stats, key=lambda x: x['pct'], reverse=True)
    for i, item in enumerate(person_stats, 1):
        ws_persons.append([
            i, item["nombres"], item["apellidos"], item["cedula"], item["cargo"],
            item["expected"], item["present"], item["absent"], f"{item['pct']}%", item["evaluation"]
        ])
    _auto_width(ws_persons)

    # ── Hoja 7: 📝 Detalle Completo de Asistencia ────────────────────────────
    ws_detail = wb.create_sheet("📝 Detalle Completo")
    ws_detail.append([
        "Proyecto", "Evento", "Fecha", "Cédula", "Nombres", "Apellidos",
        "Cargo", "Correo", "Celular", "Estado", "Tiene Firma"
    ])
    _style_header(ws_detail, 1)

    for ev in events_sorted:
        proj = ev.project
        attendance_map = {
            a.person_id: a for a in db.query(Attendance).filter(Attendance.event_id == ev.id).all()
        }
        assistants = [db.query(Person).filter(Person.id == pid).first() for pid in attendance_map.keys()]
        assistants = [p for p in assistants if p]
        for p in assistants:
            att = attendance_map.get(p.id)
            is_present = att.present if att else False
            has_signature = bool(att and att.signature) if att else False
            ws_detail.append([
                proj.name if proj else "",
                ev.name,
                ev.date,
                p.cedula,
                p.nombres,
                p.apellidos,
                p.cargo,
                p.correo or "",
                p.celular or "",
                "Presente" if is_present else "Ausente",
                "Sí" if has_signature else "No"
            ])
    _auto_width(ws_detail)

    # ── Hoja 8: 🗂️ Catálogo de Grupos ────────────────────────────────────────
    ws_groups = wb.create_sheet("🗂️ Catálogo de Grupos")
    ws_groups.append(["Grupo", "Descripción", "Total Personas", "Cédulas", "Nombres Completos"])
    _style_header(ws_groups, 1)
    for g in all_groups:
        members = g.persons
        ws_groups.append([
            g.name,
            g.description or "",
            len(members),
            ", ".join([p.cedula for p in members]),
            ", ".join([f"{p.nombres} {p.apellidos}".strip() for p in members])
        ])
    _auto_width(ws_groups)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ── Endpoint: Descargar informe Excel ─────────────────────────────────────────
@app.get("/api/reports/excel")
def download_excel_report(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    try:
        xlsx = _build_excel_report(db)
        filename = f"informe_asistencia_meta_tic_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            xlsx,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generando informe: {str(e)}")


# ── Endpoint: Estadísticas mejoradas para dashboard ──────────────────────────
@app.get("/api/stats")
def get_stats(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    # ── KPIs Básicos ──
    projects = db.query(Project).count()
    persons = db.query(Person).count()
    groups = db.query(Group).count()
    events = db.query(Event).count()
    total_att = db.query(Attendance).count()
    total_present = db.query(Attendance).filter(Attendance.present == True).count()
    attendance_pct = round((total_present / total_att * 100), 1) if total_att else 0

    # ── Eventos Recientes ──
    recent = []
    recent_evs = db.query(Event).order_by(Event.date.desc()).limit(5).all()
    for e in recent_evs:
        present = db.query(Attendance).filter(Attendance.event_id == e.id, Attendance.present == True).count()
        total_expected = db.query(Attendance).filter(Attendance.event_id == e.id).count()
        pct = round((present / total_expected * 100), 1) if total_expected else 0
        resp = None
        if e.responsible_id:
            r = db.query(Person).filter(Person.id == e.responsible_id).first()
            if r: resp = f"{r.nombres} {r.apellidos}".strip()
        recent.append({
            "id": e.id, 
            "name": e.name, 
            "date": e.date, 
            "project": e.project.name if e.project else "", 
            "present": present, 
            "expected": total_expected, 
            "pct": pct, 
            "responsible": resp
        })

    # ── Eventos Críticos (asistencia < 70%) ──
    critical_events = []
    all_events = db.query(Event).all()
    for e in all_events:
        present = db.query(Attendance).filter(Attendance.event_id == e.id, Attendance.present == True).count()
        total_expected = db.query(Attendance).filter(Attendance.event_id == e.id).count()
        if total_expected > 0:
            pct = round((present / total_expected * 100), 1)
            if pct < 70:
                critical_events.append({
                    "id": e.id,
                    "name": e.name,
                    "date": e.date,
                    "project": e.project.name if e.project else "",
                    "present": present,
                    "expected": total_expected,
                    "pct": pct
                })
    critical_events = sorted(critical_events, key=lambda x: x['pct'])[:5]  # Top 5 peores

    # ── Ranking de Proyectos ──
    project_ranking = []
    all_projects = db.query(Project).all()
    for proj in all_projects:
        proj_events = proj.events
        if not proj_events:
            continue
        total_present_proj = 0
        total_expected_proj = 0
        for ev in proj_events:
            expected_ev = db.query(Attendance).filter(Attendance.event_id == ev.id).count()
            present_ev = db.query(Attendance).filter(Attendance.event_id == ev.id, Attendance.present == True).count()
            total_present_proj += present_ev
            total_expected_proj += expected_ev
        if total_expected_proj > 0:
            pct_proj = round((total_present_proj / total_expected_proj * 100), 1)
            project_ranking.append({
                "id": proj.id,
                "name": proj.name,
                "events": len(proj_events),
                "expected": total_expected_proj,
                "present": total_present_proj,
                "pct": pct_proj
            })
    project_ranking = sorted(project_ranking, key=lambda x: x['pct'], reverse=True)
    top_projects = project_ranking[:5]
    worst_projects = project_ranking[-5:] if len(project_ranking) > 5 else []

    # ── Análisis por Grupos ──
    group_stats = []
    all_groups = db.query(Group).all()
    for g in all_groups:
        if not g.persons:
            continue
        # Contar asistencias de todos los miembros del grupo
        total_expected_group = 0
        total_present_group = 0
        for p in g.persons:
            att_person = db.query(Attendance).filter(Attendance.person_id == p.id).count()
            pres_person = db.query(Attendance).filter(Attendance.person_id == p.id, Attendance.present == True).count()
            total_expected_group += att_person
            total_present_group += pres_person
        if total_expected_group > 0:
            pct_group = round((total_present_group / total_expected_group * 100), 1)
            group_stats.append({
                "id": g.id,
                "name": g.name,
                "members": len(g.persons),
                "expected": total_expected_group,
                "present": total_present_group,
                "pct": pct_group
            })
    group_stats = sorted(group_stats, key=lambda x: x['pct'], reverse=True)

    # ── Personas con mejor y peor asistencia ──
    person_stats = []
    all_persons = db.query(Person).all()
    for p in all_persons:
        att_count = db.query(Attendance).filter(Attendance.person_id == p.id).count()
        if att_count == 0:
            continue
        pres_count = db.query(Attendance).filter(Attendance.person_id == p.id, Attendance.present == True).count()
        pct_person = round((pres_count / att_count * 100), 1) if att_count else 0
        person_stats.append({
            "id": p.id,
            "name": f"{p.nombres} {p.apellidos}".strip(),
            "cedula": p.cedula,
            "cargo": p.cargo,
            "expected": att_count,
            "present": pres_count,
            "pct": pct_person
        })
    person_stats = sorted(person_stats, key=lambda x: x['pct'], reverse=True)
    top_persons = person_stats[:5]
    worst_persons = person_stats[-5:] if len(person_stats) > 5 else []

    # ── Tendencia de asistencia (últimos 10 eventos) ──
    trend_events = db.query(Event).order_by(Event.date.desc()).limit(10).all()
    trend_data = []
    for e in reversed(trend_events):  # Orden cronológico
        present = db.query(Attendance).filter(Attendance.event_id == e.id, Attendance.present == True).count()
        total_expected = db.query(Attendance).filter(Attendance.event_id == e.id).count()
        pct = round((present / total_expected * 100), 1) if total_expected else 0
        trend_data.append({
            "date": e.date,
            "name": e.name,
            "pct": pct
        })

    return {
        # KPIs básicos
        "projects": projects,
        "persons": persons,
        "groups": groups,
        "events": events,
        "attendance_records": total_att,
        "attendance_present": total_present,
        "attendance_pct": attendance_pct,
        
        # Eventos
        "recent_events": recent,
        "critical_events": critical_events,
        
        # Rankings
        "top_projects": top_projects,
        "worst_projects": worst_projects,
        "top_persons": top_persons,
        "worst_persons": worst_persons,
        
        # Análisis
        "group_stats": group_stats,
        "trend_data": trend_data
    }


# ── Endpoints: Backup / Restore de la base de datos (descargar / subir) ──────
@app.get("/api/db/download")
def download_db(current_user: User = Depends(get_current_user)):
    db_file = resolve_sqlite_path()
    if not os.path.exists(db_file):
        raise HTTPException(404, "Archivo de base de datos no encontrado")
    return FileResponse(path=db_file, filename=os.path.basename(db_file), media_type='application/octet-stream')


@app.post("/api/db/upload")
def upload_db(file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    db_file = resolve_sqlite_path()
    tmp_path = db_file + ".upload.tmp"
    try:
        # Guardar carga temporal
        with open(tmp_path, 'wb') as out_f:
            shutil.copyfileobj(file.file, out_f)

        # Preparar backup path variable
        bak_path = None

        # Hacer backup del actual
        if os.path.exists(db_file):
            bak_path = f"{db_file}.bak-{int(time.time())}"
            shutil.copy2(db_file, bak_path)

        # Reemplazar
        try:
            # Declarar globals antes de usar engine
            global engine, SessionLocal
            # Cerrar conexiones activas antes de reemplazar
            try:
                engine.dispose()
            except Exception:
                pass
            os.replace(tmp_path, db_file)
        except Exception as e:
            raise HTTPException(500, f"No se pudo reemplazar la base de datos: {str(e)}")

        # Recrear engine y session local para usar la nueva base
        engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

        # Asegurar que esquema esperado exista
        try:
            Base.metadata.create_all(bind=engine, checkfirst=True)
        except Exception:
            pass

        return {"ok": True, "message": "Base de datos restaurada. Backup creado.", "backup": bak_path if os.path.exists(db_file) else None}
    finally:
        try:
            file.file.close()
        except Exception:
            pass

# ── Servir frontend estático ──────────────────────────────────────────────────
frontend_path = os.path.join(os.path.dirname(__file__), "..", "frontend")
print(f"🔍 Buscando frontend en: {frontend_path}")
print(f"🔍 Frontend existe: {os.path.exists(frontend_path)}")

if os.path.exists(frontend_path):
    print(f"✅ Montando frontend desde {frontend_path}")
    app.mount("/", StaticFiles(directory=frontend_path, html=True), name="frontend")
else:
    print(f"⚠️  Advertencia: No se encontró el directorio frontend en {frontend_path}")
    
    # Fallback: servir un mensaje simple en la raíz
    @app.get("/")
    def root():
        return {
            "app": "Asistencia API",
            "version": "1.0.0",
            "status": "running",
            "docs": "/docs",
            "api": "/api"
        }

